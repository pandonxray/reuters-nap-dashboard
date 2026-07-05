from __future__ import annotations

import hashlib
import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

logger = logging.getLogger(__name__)

EXCEL_EPOCH = "1899-12-30"
GALLONS_PER_PETROLEUM_BARREL = 42.0
NAPHTHA_BBLS_PER_METRIC_TON = 8.9
EBOB_BBLS_PER_METRIC_TON = 8.33
GASOIL_BBLS_PER_METRIC_TON = 7.45
DEFAULT_NAP_WORKBOOK = Path(
    r"C:\Users\74100\Nutstore\1\油气-djx-\NAP-丙烯-坚果云\Nap_calendar_month_live_formula.xlsx"
)
STANDARD_COLUMNS = [
    "date",
    "series_id",
    "display_name",
    "value",
    "sheet",
    "sector",
    "product",
    "region",
    "contract_month",
    "term_type",
    "calendar_month",
    "unit_native",
    "unit_normalized",
    "ric",
    "is_derived",
    "source",
]
EXTRA_COLUMNS = ["value_normalized", "unit_conversion", "unit_source", "name_native", "short_name"]
RELEVANT_SHEETS = [
    "Crude",
    "Gasoline",
    "Heating Oil&Jet fuel",
    "Diesel",
    "Nap",
    "LNG",
    "Crk",
    "Margin",
    "Propane",
    "Fuel oil",
    "Freight",
    "原油",
    "成品油(国内汽柴表)",
    "成品油(国外汽柴表)",
]

CALENDAR_MONTH_SHEETS = {
    "Gasoline_自然月": "Gasoline",
    "HOJet_自然月": "Heating Oil&Jet fuel",
    "Diesel_自然月": "Diesel",
    "Nap_自然月": "Nap",
    "LNG_自然月": "LNG",
    "Crk_自然月": "Crk",
    "Margin_自然月": "Margin",
    "Propane_自然月": "Propane",
    "FuelOil_自然月": "Fuel oil",
}
CALENDAR_OUTPUT_BY_SOURCE = {source: calendar for calendar, source in CALENDAR_MONTH_SHEETS.items()}

CALENDAR_LABEL_ALIASES = {
    "RB": "RBOB",
    "EBOBNWE": "EBOB",
    "MOG92SG": "新加坡92汽油纸货",
    "MOG95SG": "新加坡95汽油纸货",
    "HO": "HO",
    "JETFUSGC": "Jet USG",
    "JETFCNWE": "Jet NWE",
    "JETSG": "Jet Singapore",
    "LGO": "LSGO",
    "GO10SG": "新加坡10ppm柴油纸货",
    "NACFRJP": "MOPJ",
    "NAPJPEW": "石脑油东西价差",
    "NACFRJPCK": "MOPJ裂差",
    "NAPCNWEA": "NWE Naphtha",
    "NAPCNWEAC": "NWE Naphtha Crack",
    "A7Q": "美国天然气",
    "MOG92SGCK": "新加坡92汽油裂差",
    "EBOBNWECK": "EBOB裂差",
    "RBCCLC": "RBOB裂差",
    "GO10BRTCK": "新加坡柴油裂差",
    "LGOC": "欧洲柴油裂差",
    "HOCCLC": "HO裂差",
    "JETSGCK": "新加坡航煤裂差",
    "JETFCNWECK": "欧洲航煤裂差",
    "FO380BRTCK": "新加坡高硫裂差",
    "HFOFARAAC": "欧洲高硫裂差",
    "PROCNWE": "NWE LPG",
    "PROFEI": "FEI",
    "FO380SG": "新加坡高硫燃料油",
    "HFOFARAA": "欧洲高硫燃料油",
}

SECTOR_BY_SHEET = {
    "Crude": "Crude",
    "Gasoline": "Gasoline",
    "Heating Oil&Jet fuel": "Jet/Heating Oil",
    "Diesel": "Diesel",
    "Nap": "Naphtha",
    "LNG": "LNG",
    "Crk": "Cracks",
    "Margin": "Margins",
    "Propane": "Propane/LPG",
    "Fuel oil": "Fuel Oil",
    "Freight": "Freight",
    "原油": "Freight",
    "成品油(国内汽柴表)": "Freight",
    "成品油(国外汽柴表)": "Freight",
}

GENERIC_META_LABELS = {
    "序列",
    "数据模块",
    "数据简称",
    "数据名称",
    "路透代码",
    "timestamp",
    "voyage rate (bbl)",
}


@dataclass(frozen=True)
class ParsedSeries:
    sheet: str
    display_name: str
    short_name: str
    name_native: str
    ric: str
    source: str
    is_derived: bool
    date_col: int
    value_col: int
    header_row: int
    first_data_row: int


def project_root() -> Path:
    return Path(__file__).resolve().parents[1]


def default_cache_path() -> Path:
    return project_root() / "data" / "processed" / "nap_timeseries.parquet"


def default_catalog_path() -> Path:
    return project_root() / "config" / "nap_series_catalog.yaml"


def default_explanations_path() -> Path:
    return project_root() / "config" / "nap_explanations.yaml"


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and np.isnan(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""
    return text


def _norm(value: Any) -> str:
    return re.sub(r"\s+", "", _clean_text(value)).lower()


def _hash_text(text: str) -> str:
    return hashlib.blake2b(text.encode("utf-8"), digest_size=5).hexdigest()


def _slug(value: str) -> str:
    ascii_text = re.sub(r"[^A-Za-z0-9]+", "_", value).strip("_").lower()
    ascii_text = re.sub(r"_+", "_", ascii_text)
    return ascii_text


def _make_series_id(sheet: str, display_name: str, ric: str, value_col: int) -> str:
    basis = ric or display_name or f"col_{value_col}"
    slug = _slug(f"{sheet}_{basis}")
    if not slug:
        slug = f"{_slug(sheet)}_{_hash_text(basis) or value_col}"
    suffix = _hash_text(f"{sheet}|{display_name}|{ric}|{value_col}")
    return f"{slug}_{suffix}"


def _continuous_ric_base(ric: str) -> str:
    value = str(ric or "").upper()
    value = re.sub(r"(?:SW)?MC\d{1,2}$", "", value)
    if re.search(r"[-=/]", value):
        value = re.sub(r"C\d{1,2}(?=$|[-=/])", "C", value)
    else:
        value = re.sub(r"C\d{1,2}$", "", value)
    return re.sub(r"[^A-Z0-9]+", "", value)


def _calendar_base_label(display_name: str, ricbase: str) -> str:
    if ricbase in CALENDAR_LABEL_ALIASES:
        return CALENDAR_LABEL_ALIASES[ricbase]
    label = display_name or ricbase
    for pattern in [
        r"\b(monthly\s+)?continuation\s*\d{1,2}\b",
        r"\bmonth\s+continuation\s*\d{1,2}\b",
        r"\bM\s*\d{1,2}\b",
        r"\bc\s*\d{1,2}\b",
        r"连\s*\d{1,2}",
    ]:
        label = re.sub(pattern, "", label, flags=re.IGNORECASE)
    label = re.sub(r"\s+", " ", label).strip(" -_/")
    return label or ricbase


def usd_per_gallon_to_usd_per_barrel(value: float | pd.Series) -> float | pd.Series:
    return value * GALLONS_PER_PETROLEUM_BARREL


def usd_per_metric_ton_to_usd_per_barrel(value: float | pd.Series, barrels_per_metric_ton: float) -> float | pd.Series:
    return value / barrels_per_metric_ton


def _unit_rule(sheet: str, display_name: str, ric: str) -> dict[str, Any]:
    text = f"{display_name} {ric} {sheet}".lower()
    sheet_text = str(sheet or "").lower()
    ric_upper = str(ric or "").upper()

    if (
        "crack" in text
        or "crk" in sheet_text
        or sheet == "Crk"
        or ric_upper.startswith(("NAPCNWEAC", "CAL_NAPCNWEAC"))
        or "CK" in ric_upper
    ):
        return {
            "unit_native": "USD/bbl",
            "unit_normalized": "USD/bbl",
            "factor": 1.0,
            "conversion": "原始已为 USD/bbl，价差/裂解计算不做桶吨换算。",
            "source": "Reuters crack spread / exchange crack-spread convention",
        }
    if ric_upper.startswith(("RB", "HO")) or "rbob" in text or re.search(r"\bho\b", text):
        return {
            "unit_native": "USD/gal",
            "unit_normalized": "USD/bbl",
            "factor": GALLONS_PER_PETROLEUM_BARREL,
            "conversion": f"USD/gal × {GALLONS_PER_PETROLEUM_BARREL:.0f} = USD/bbl。",
            "source": "CME NYMEX refined products quote convention",
        }
    if ric_upper.startswith(("EBOBNWE", "CAL_EBOBNWE")) or "ebob" in text:
        return {
            "unit_native": "USD/mt",
            "unit_normalized": "USD/bbl",
            "factor": 1.0 / EBOB_BBLS_PER_METRIC_TON,
            "conversion": f"EBOB: USD/mt ÷ {EBOB_BBLS_PER_METRIC_TON:.2f} = USD/bbl。",
            "source": "ICE/CME Eurobob contract conventions; 1,000 mt = 8,330 bbl",
        }
    if ric_upper.startswith(("NAPCNWEAM", "CAL_NAPCNWEA", "NACFRJP", "CAL_NACFRJP")) or (
        "naphtha" in text and "crack" not in text
    ):
        return {
            "unit_native": "USD/mt",
            "unit_normalized": "USD/bbl",
            "factor": 1.0 / NAPHTHA_BBLS_PER_METRIC_TON,
            "conversion": f"石脑油: USD/mt ÷ {NAPHTHA_BBLS_PER_METRIC_TON:.2f} = USD/bbl。",
            "source": "Naphtha outright quote convention plus dashboard barrel conversion",
        }
    if ric_upper.startswith(("LGO", "CAL_LGO")) or "gasoil" in text or "lsgo" in text:
        return {
            "unit_native": "USD/mt",
            "unit_normalized": "USD/bbl",
            "factor": 1.0 / GASOIL_BBLS_PER_METRIC_TON,
            "conversion": f"LSGO/Gasoil: USD/mt ÷ {GASOIL_BBLS_PER_METRIC_TON:.2f} = USD/bbl。",
            "source": "ICE gasoil crack-spread convention",
        }
    if sheet == "Crude" or any(token in text for token in ["wti", "brent", "dubai", "oman"]):
        return {
            "unit_native": "USD/bbl",
            "unit_normalized": "USD/bbl",
            "factor": 1.0,
            "conversion": "原始已为 USD/bbl。",
            "source": "Crude futures / swaps quote convention",
        }
    if sheet in {"Crk", "Margin"}:
        return {
            "unit_native": "USD/bbl",
            "unit_normalized": "USD/bbl",
            "factor": 1.0,
            "conversion": "原始已为 USD/bbl。",
            "source": "Reuters crack/margin convention",
        }
    if sheet in {"Freight", "原油", "成品油(国内汽柴表)", "成品油(国外汽柴表)"}:
        return {
            "unit_native": "USD/bbl",
            "unit_normalized": "USD/bbl",
            "factor": 1.0,
            "conversion": "原始已为 USD/bbl。",
            "source": "Reuters freight route convention",
        }
    if sheet in {"Diesel", "Nap", "Propane", "Fuel oil"}:
        return {
            "unit_native": "USD/mt",
            "unit_normalized": "USD/mt",
            "factor": 1.0,
            "conversion": "原始为 USD/mt；未参与跨桶价价差时保留吨价。",
            "source": "Reuters refined products quote convention",
        }
    if sheet == "Gasoline":
        return {
            "unit_native": "USD/bbl",
            "unit_normalized": "USD/bbl",
            "factor": 1.0,
            "conversion": "原始已为 USD/bbl。",
            "source": "Reuters gasoline quote convention",
        }
    if sheet == "LNG":
        return {
            "unit_native": "USD/MMBtu",
            "unit_normalized": "USD/MMBtu",
            "factor": 1.0,
            "conversion": "原始已为 USD/MMBtu。",
            "source": "Reuters LNG/gas quote convention",
        }
    return {"unit_native": "", "unit_normalized": "", "factor": 1.0, "conversion": "", "source": ""}


def _apply_unit_rule_metadata(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame
    out = frame.copy()
    for col in ["unit_native", "unit_normalized", "unit_conversion", "unit_source"]:
        if col not in out.columns:
            out[col] = ""
    if "value_normalized" not in out.columns:
        out["value_normalized"] = out.get("value", np.nan)
    key_col = "series_id" if "series_id" in out.columns else None
    meta_cols = [col for col in ["series_id", "sheet", "display_name", "ric"] if col in out.columns]
    if not meta_cols:
        return out
    meta = out[meta_cols].drop_duplicates(key_col or meta_cols)
    rule_rows: list[dict[str, Any]] = []
    for row in meta.itertuples(index=False):
        row_dict = dict(zip(meta_cols, row, strict=False))
        rule = _unit_rule(str(row_dict.get("sheet", "")), str(row_dict.get("display_name", "")), str(row_dict.get("ric", "")))
        if not rule.get("unit_native"):
            continue
        rule_rows.append(
            {
                "key": str(row_dict.get(key_col, len(rule_rows))) if key_col else "|".join(str(row_dict.get(col, "")) for col in meta_cols),
                "unit_native": rule["unit_native"],
                "unit_normalized": rule["unit_normalized"],
                "unit_conversion": rule["conversion"],
                "unit_source": rule["source"],
                "factor": float(rule["factor"]),
            }
        )
    if not rule_rows:
        return out
    rules = pd.DataFrame(rule_rows).drop_duplicates("key").set_index("key")
    keys = out[key_col].astype(str) if key_col else out[meta_cols].astype(str).agg("|".join, axis=1)
    for col in ["unit_native", "unit_normalized", "unit_conversion", "unit_source"]:
        mapped = keys.map(rules[col])
        mask = mapped.notna()
        if mask.any():
            out.loc[mask, col] = mapped[mask].astype(str).to_numpy()
    factor = keys.map(rules["factor"]).fillna(1.0).astype(float)
    out["value_normalized"] = pd.to_numeric(out["value"], errors="coerce") * factor.to_numpy()
    return out


def coerce_excel_dates(values: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(values, errors="coerce")
    parsed = pd.to_datetime(values, errors="coerce")
    serial_mask = numeric.notna() & numeric.between(15000, 90000)
    if serial_mask.any():
        parsed.loc[serial_mask] = pd.to_datetime(
            numeric.loc[serial_mask],
            unit="D",
            origin=EXCEL_EPOCH,
            errors="coerce",
        )
    return parsed


def _looks_like_date_series(values: pd.Series, min_points: int = 8) -> bool:
    return int(coerce_excel_dates(values).notna().sum()) >= min_points


def _looks_like_numeric_series(values: pd.Series, min_points: int = 8) -> bool:
    return int(pd.to_numeric(values, errors="coerce").notna().sum()) >= min_points


def _worksheet_to_frame(ws) -> pd.DataFrame:
    return pd.DataFrame(ws.iter_rows(values_only=True))


def _top_reuters_formulas(ws, max_rows: int = 8) -> list[tuple[int, int, str, str]]:
    formulas: list[tuple[int, int, str, str]] = []
    for row in ws.iter_rows(min_row=1, max_row=min(max_rows, ws.max_row)):
        for cell in row:
            value = cell.value
            if not isinstance(value, str) or not value.startswith("="):
                continue
            upper = value.upper()
            if "_XLL.RDP.HISTORICALPRICING" in upper:
                formulas.append((cell.row, cell.column, "RDP.HistoricalPricing", value))
            elif "_XLL.RHISTORY" in upper:
                formulas.append((cell.row, cell.column, "RHistory", value))
    return formulas


def _formula_output_anchor(formula: str) -> tuple[int, int] | None:
    match = re.search(r",\s*([A-Z]{1,3})(\d+)\s*\)\s*$", formula)
    if not match:
        return None
    return column_index_from_string(match.group(1)), int(match.group(2))


def _first_non_empty_above(df: pd.DataFrame, row0: int, col0: int) -> str:
    for row in range(max(row0 - 1, 0), -1, -1):
        value = _clean_text(df.iat[row, col0]) if col0 < df.shape[1] else ""
        if value and _norm(value) not in GENERIC_META_LABELS and not value.startswith("="):
            return value
    return ""


def _candidate_meta_rows(df: pd.DataFrame, max_left_col0: int) -> list[int]:
    rows: list[int] = []
    max_scan = min(df.shape[0], 1200)
    for row in range(max_scan):
        values = [_clean_text(df.iat[row, col]) for col in range(max(0, max_left_col0))]
        useful = [
            value
            for value in values
            if value
            and not value.startswith("=")
            and _norm(value) not in GENERIC_META_LABELS
            and not re.fullmatch(r"\d+(\.0)?", value)
        ]
        if useful:
            rows.append(row)
    return rows


def _looks_like_ric(value: str) -> bool:
    if not value or any("\u4e00" <= char <= "\u9fff" for char in value):
        return False
    if " " in value or len(value) < 3 or len(value) > 40:
        return False
    if not re.search(r"[A-Za-z]", value):
        return False
    return bool(re.fullmatch(r"[A-Za-z0-9.\-=/#]+", value))


def _extract_ric_from_formula(formula: str) -> list[str]:
    match = re.search(r'RHistory\("([^"]+)"', formula, flags=re.IGNORECASE)
    if not match:
        return []
    return [part.strip() for part in match.group(1).split(";") if part.strip()]


def _metadata_from_row(df: pd.DataFrame, row: int | None, max_left_col0: int, header: str) -> dict[str, str]:
    if row is None or row < 0 or row >= df.shape[0]:
        return {"ric": "", "short_name": "", "name_native": ""}
    cells = [_clean_text(df.iat[row, col]) for col in range(max(0, max_left_col0))]
    texts = [
        value
        for value in cells
        if value
        and not value.startswith("=")
        and _norm(value) not in GENERIC_META_LABELS
        and not re.fullmatch(r"\d+(\.0)?", value)
    ]
    ric_candidates = [value for value in texts if _looks_like_ric(value)]
    ric = ric_candidates[-1] if ric_candidates else ""
    text_candidates = [value for value in texts if value != ric]
    header_norm = _norm(header)
    exact = [value for value in text_candidates if _norm(value) == header_norm]
    short_name = exact[0] if exact else ""
    if not short_name:
        short_candidates = [value for value in text_candidates if re.search(r"[A-Za-z0-9]", value) and len(value) <= 40]
        short_name = short_candidates[-1] if short_candidates else ""
    native_candidates = [value for value in text_candidates if value != short_name]
    name_native = native_candidates[0] if native_candidates else (short_name or header)
    return {"ric": ric, "short_name": short_name, "name_native": name_native}


def _find_meta_row(df: pd.DataFrame, max_left_col0: int, header: str, pair_index: int) -> int | None:
    header_norm = _norm(header)
    rows = _candidate_meta_rows(df, max_left_col0)
    return _find_meta_row_from_candidates(df, max_left_col0, header_norm, pair_index, rows)


def _find_meta_row_from_candidates(
    df: pd.DataFrame,
    max_left_col0: int,
    header_norm: str,
    pair_index: int,
    rows: list[int],
    left_norm_values: dict[int, list[str]] | None = None,
) -> int | None:
    if header_norm:
        for row in rows:
            left_values = (
                left_norm_values[row]
                if left_norm_values is not None and row in left_norm_values
                else [_norm(df.iat[row, col]) for col in range(max(0, max_left_col0))]
            )
            if header_norm in left_values:
                return row
        for row in rows:
            left_values = (
                left_norm_values[row]
                if left_norm_values is not None and row in left_norm_values
                else [_norm(df.iat[row, col]) for col in range(max(0, max_left_col0))]
            )
            if any(header_norm and (header_norm in value or value in header_norm) for value in left_values if value):
                return row
    if pair_index < len(rows):
        return rows[pair_index]
    return None


def _find_pair_start_near(df: pd.DataFrame, anchor_row0: int, anchor_col0: int) -> tuple[int, int]:
    for offset in range(0, 8):
        col = anchor_col0 + offset
        if col + 1 >= df.shape[1]:
            break
        date_values = df.iloc[anchor_row0:, col]
        value_values = df.iloc[anchor_row0:, col + 1]
        if _looks_like_date_series(date_values) and _looks_like_numeric_series(value_values):
            return anchor_row0, col
    # Some formula anchors point at the header row rather than the first data row.
    for row_offset in range(1, 4):
        row = anchor_row0 + row_offset
        if row >= df.shape[0]:
            break
        for offset in range(0, 8):
            col = anchor_col0 + offset
            if col + 1 >= df.shape[1]:
                break
            if _looks_like_date_series(df.iloc[row:, col]) and _looks_like_numeric_series(df.iloc[row:, col + 1]):
                return row, col
    return anchor_row0, anchor_col0


def _infer_sector(sheet: str, display_name: str, ric: str = "") -> str:
    text = f"{display_name} {ric} {sheet}".lower()
    ric_upper = ric.upper()
    if sheet == "Nap" and (
        "crack spread" in text
        or ric_upper.startswith("NAPCNWEAC")
        or ric_upper.startswith("NACFRJPCK")
    ):
        return "Cracks"
    return SECTOR_BY_SHEET.get(sheet, "Other")


def _infer_region(display_name: str, ric: str, sheet: str) -> str:
    text = f"{display_name} {ric} {sheet}".lower()
    if any(token in text for token in ["singapore", "新加坡", "sin"]):
        return "Singapore"
    if any(token in text for token in ["china", "中国", "宁波", "青岛", "tao"]):
        return "China"
    if any(token in text for token in ["japan", "日本", "tokyo", "tyo"]):
        return "Japan"
    if any(token in text for token in ["korea", "韩国", "yos"]):
        return "Korea"
    if any(token in text for token in ["usgc", "美湾", "hou", "美国", "wti", "nymex"]):
        return "US"
    if "nwe" in text:
        return "Northwest Europe"
    if any(token in text for token in ["europe", "欧洲", "nwe", "rotterdam", "阿姆斯特丹", "france", "法国"]):
        return "Europe"
    if any(token in text for token in ["middle east", "中东", "ras tanura", "dubai"]):
        return "Middle East"
    if any(token in text for token in ["med", "地中海", "lavera"]):
        return "Mediterranean"
    return "Global"


def _infer_product(display_name: str, ric: str, sheet: str) -> str:
    text = f"{display_name} {ric} {sheet}".lower()
    ric_upper = ric.upper()
    if ric_upper.startswith("NAPCNWEAM") or "naphtha cif nwe outright" in text:
        return "NWE CIF Naphtha"
    if ric_upper.startswith("NAPCNWEAC") or "naphtha cif nwe" in text and "crack" in text:
        return "NWE Naphtha Crack"
    if ric_upper.startswith("NACFRJPCK") or "mopj" in text and "crack" in text:
        return "MOPJ Crack"
    if "wti" in text or ric.startswith("CL"):
        return "WTI"
    if "brent" in text or "brt" in text or ric.startswith("LCO"):
        return "Brent"
    if "dubai" in text:
        return "Dubai"
    if "rbob" in text or ric.startswith("RB"):
        return "RBOB"
    if re.search(r"\bho\b", text) or ric.startswith("HO"):
        return "Heating Oil"
    if "gasoil" in text or ric.startswith("LGO"):
        return "LSGO"
    if "92" in text and "crack" in text:
        return "Singapore 92R Crack"
    if "crack" in text:
        return "Crack"
    if "margin" in text or "炼厂" in text:
        return "Refining Margin"
    if "vlsfo" in text or "低硫" in text:
        return "VLSFO"
    if "hsfo" in text or "高硫" in text or "fo380" in text:
        return "HSFO"
    if "fuel oil" in text or "燃料油" in text:
        return "Fuel Oil"
    if "nap" in text or "naf" in text or "石脑油" in text or "naphtha" in text:
        return "Naphtha"
    if "propane" in text or "lpg" in text or "丙烷" in text:
        return "Propane"
    if sheet in {"原油", "成品油(国内汽柴表)", "成品油(国外汽柴表)"}:
        return "Freight"
    return SECTOR_BY_SHEET.get(sheet, sheet)


def _infer_contract_month(display_name: str, ric: str) -> str:
    text = f"{display_name} {ric}"
    for pattern in [r"\bM(\d{1,2})\b", r"连\s*(\d{1,2})", r"c(\d{1,2})\b"]:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return f"M{int(match.group(1))}"
    return ""


def _infer_unit(sheet: str, display_name: str, ric: str) -> tuple[str, str]:
    rule = _unit_rule(sheet, display_name, ric)
    return str(rule["unit_native"]), str(rule["unit_normalized"])


def _series_to_records(
    df: pd.DataFrame,
    parsed: ParsedSeries,
    catalog_override: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    date_values = coerce_excel_dates(df.iloc[parsed.first_data_row :, parsed.date_col])
    values = pd.to_numeric(df.iloc[parsed.first_data_row :, parsed.value_col], errors="coerce")
    frame = pd.DataFrame({"date": date_values, "value": values}).dropna(subset=["date", "value"])
    if frame.empty:
        return []

    display_name = parsed.display_name or parsed.short_name or parsed.name_native or parsed.ric
    series_id = _make_series_id(parsed.sheet, display_name, parsed.ric, parsed.value_col)
    unit_native, unit_normalized = _infer_unit(parsed.sheet, display_name, parsed.ric)
    sector = _infer_sector(parsed.sheet, display_name, parsed.ric)
    product = _infer_product(display_name, parsed.ric, parsed.sheet)
    region = _infer_region(display_name, parsed.ric, parsed.sheet)
    contract_month = _infer_contract_month(display_name, parsed.ric)

    if catalog_override:
        display_name = catalog_override.get("display_name") or display_name
        sector = catalog_override.get("sector") or sector
        product = catalog_override.get("product") or product
        region = catalog_override.get("region") or region
        contract_month = catalog_override.get("contract_month") or contract_month
        unit_native = catalog_override.get("unit_native") or unit_native
        unit_normalized = catalog_override.get("unit_normalized") or unit_normalized

    unit_rule = _unit_rule(parsed.sheet, display_name, parsed.ric)
    if unit_rule.get("unit_native"):
        unit_native = str(unit_rule["unit_native"])
        unit_normalized = str(unit_rule["unit_normalized"])
    frame["value_normalized"] = frame["value"] * float(unit_rule.get("factor", 1.0))

    rows: list[dict[str, Any]] = []
    for row in frame.itertuples(index=False):
        rows.append(
            {
                "date": pd.Timestamp(row.date).normalize(),
                "series_id": series_id,
                "display_name": display_name,
                "value": float(row.value),
                "sheet": parsed.sheet,
                "sector": sector,
                "product": product,
                "region": region,
                "contract_month": contract_month,
                "term_type": "continuous",
                "calendar_month": "",
                "unit_native": unit_native,
                "unit_normalized": unit_normalized,
                "ric": parsed.ric,
                "is_derived": bool(parsed.is_derived),
                "source": parsed.source,
                "value_normalized": float(row.value_normalized),
                "unit_conversion": str(unit_rule.get("conversion", "")),
                "unit_source": str(unit_rule.get("source", "")),
                "name_native": parsed.name_native,
                "short_name": parsed.short_name,
            }
        )
    return rows


def _contract_month_number(value: Any) -> int | None:
    match = re.search(r"(\d{1,2})", str(value or ""))
    if not match:
        return None
    number = int(match.group(1))
    return number if 1 <= number <= 12 else None


def _derive_calendar_month_records(continuous: pd.DataFrame) -> list[dict[str, Any]]:
    if continuous.empty:
        return []
    frame = continuous[continuous.get("term_type", "continuous").eq("continuous")].copy()
    if frame.empty:
        return []
    frame["month_num"] = frame["contract_month"].map(_contract_month_number)
    frame = frame[frame["month_num"].between(1, 12, inclusive="both")]
    frame = frame[frame["sheet"].isin(set(CALENDAR_OUTPUT_BY_SOURCE))]
    if frame.empty:
        return []

    catalog = (
        frame.sort_values(["sheet", "product", "region", "contract_month", "display_name"])
        .drop_duplicates("series_id")
        .copy()
    )
    catalog["ricbase"] = catalog["ric"].map(_continuous_ric_base)
    catalog = catalog[catalog["ricbase"].astype(bool)]

    records: list[dict[str, Any]] = []
    for (source_sheet, ricbase), group in catalog.groupby(["sheet", "ricbase"], dropna=False):
        month_catalog = group.drop_duplicates("month_num")
        if set(month_catalog["month_num"].astype(int)) != set(range(1, 13)):
            continue
        month_catalog = month_catalog.sort_values("month_num")
        series_ids = list(month_catalog["series_id"])
        data = frame[frame["series_id"].isin(series_ids)].copy()
        if data.empty:
            continue
        raw_wide = data.pivot_table(index="date", columns="month_num", values="value", aggfunc="last").sort_index()
        norm_wide = data.pivot_table(index="date", columns="month_num", values="value_normalized", aggfunc="last").sort_index()
        if raw_wide.empty:
            continue

        meta = month_catalog[month_catalog["month_num"] == 1].iloc[0].to_dict()
        base_label = _calendar_base_label(str(meta.get("display_name") or ""), str(ricbase))
        calendar_sheet = CALENDAR_OUTPUT_BY_SOURCE.get(str(source_sheet), f"{source_sheet}_自然月")
        calendar_sector = str(meta.get("sector") or _infer_sector(str(source_sheet), base_label, str(meta.get("ric") or "")))
        calendar_product = str(meta.get("product") or _infer_product(base_label, str(meta.get("ric") or ""), str(source_sheet)))
        calendar_region = str(meta.get("region") or _infer_region(base_label, str(meta.get("ric") or ""), str(source_sheet)))
        unit_native = str(meta.get("unit_native") or "")
        unit_normalized = str(meta.get("unit_normalized") or unit_native)
        unit_conversion = str(meta.get("unit_conversion") or "")
        unit_source = str(meta.get("unit_source") or "")

        month_by_date = pd.Series(raw_wide.index.month, index=raw_wide.index)
        for target_month in range(1, 13):
            selected_contract = ((target_month - month_by_date) % 12) + 1
            raw_values = pd.Series(np.nan, index=raw_wide.index, dtype=float)
            norm_values = pd.Series(np.nan, index=raw_wide.index, dtype=float)
            for contract_num in range(1, 13):
                mask = selected_contract == contract_num
                if contract_num in raw_wide.columns:
                    raw_values.loc[mask] = raw_wide.loc[mask, contract_num]
                if contract_num in norm_wide.columns:
                    norm_values.loc[mask] = norm_wide.loc[mask, contract_num]
            valid = raw_values.dropna()
            if valid.empty:
                continue

            display_name = f"{base_label} {target_month}月"
            ric = f"CAL_{ricbase}_{target_month:02d}"
            series_id = _make_series_id(calendar_sheet, display_name, ric, target_month)
            for date, value in valid.items():
                normalized = norm_values.get(date, np.nan)
                if pd.isna(normalized):
                    normalized = value
                records.append(
                    {
                        "date": pd.Timestamp(date).normalize(),
                        "series_id": series_id,
                        "display_name": display_name,
                        "value": float(value),
                        "sheet": calendar_sheet,
                        "sector": calendar_sector,
                        "product": calendar_product,
                        "region": calendar_region,
                        "contract_month": "",
                        "term_type": "calendar",
                        "calendar_month": int(target_month),
                        "unit_native": unit_native,
                        "unit_normalized": unit_normalized,
                        "ric": ric,
                        "is_derived": True,
                        "source": "CalendarMonthFormula",
                        "value_normalized": float(normalized),
                        "unit_conversion": unit_conversion,
                        "unit_source": unit_source,
                        "name_native": display_name,
                        "short_name": display_name,
                    }
                )
    return records


def _parse_rdp_sheet(sheet: str, df: pd.DataFrame, formula: tuple[int, int, str, str]) -> list[ParsedSeries]:
    _, _, source, formula_text = formula
    anchor = _formula_output_anchor(formula_text)
    if anchor is None:
        anchor_col0, anchor_row0 = 0, 1
    else:
        anchor_col0, anchor_row0 = anchor[0] - 1, anchor[1] - 1
    first_data_row, first_date_col = _find_pair_start_near(df, anchor_row0, anchor_col0)

    parsed: list[ParsedSeries] = []
    invalid_streak = 0
    pair_index = 0
    meta_rows = _candidate_meta_rows(df, first_date_col)
    left_norm_values = {
        row: [_norm(df.iat[row, col]) for col in range(max(0, first_date_col))]
        for row in meta_rows
    }
    for date_col in range(first_date_col, df.shape[1] - 1, 2):
        value_col = date_col + 1
        date_series = df.iloc[first_data_row:, date_col]
        value_series = df.iloc[first_data_row:, value_col]
        valid_dates = coerce_excel_dates(date_series).notna()
        valid_values = pd.to_numeric(value_series, errors="coerce").notna()
        valid_count = int((valid_dates & valid_values).sum())
        if valid_count < 8:
            invalid_streak += 1
            if invalid_streak >= 3:
                break
            continue
        invalid_streak = 0
        header = _first_non_empty_above(df, first_data_row, value_col)
        meta_row = _find_meta_row_from_candidates(
            df,
            first_date_col,
            _norm(header),
            pair_index,
            meta_rows,
            left_norm_values,
        )
        meta = _metadata_from_row(df, meta_row, first_date_col, header)
        display_name = header or meta["short_name"] or meta["name_native"] or meta["ric"] or f"{sheet} {pair_index + 1}"
        parsed.append(
            ParsedSeries(
                sheet=sheet,
                display_name=display_name,
                short_name=meta["short_name"],
                name_native=meta["name_native"],
                ric=meta["ric"],
                source=source,
                is_derived=False,
                date_col=date_col,
                value_col=value_col,
                header_row=max(first_data_row - 1, 0),
                first_data_row=first_data_row,
            )
        )
        pair_index += 1
    return parsed


def _parse_rhistory_formula_group(sheet: str, df: pd.DataFrame, formula: tuple[int, int, str, str]) -> list[ParsedSeries]:
    row, col, source, formula_text = formula
    col0 = col - 1
    date_col = col0 + 1
    if date_col >= df.shape[1]:
        return []

    header_row0 = row if row < df.shape[0] else max(row - 1, 0)
    first_data_row = header_row0 + 1
    if not _looks_like_date_series(df.iloc[first_data_row:, date_col], min_points=5):
        # RHistory often has formula on row 1/2, field headers one row below, data below that.
        for candidate_row in range(max(0, row), min(df.shape[0] - 1, row + 4)):
            if _looks_like_date_series(df.iloc[candidate_row + 1 :, date_col], min_points=5):
                header_row0 = candidate_row
                first_data_row = candidate_row + 1
                break

    ric_list = _extract_ric_from_formula(formula_text)
    parsed: list[ParsedSeries] = []
    value_col = date_col + 1
    ric_index = 0
    while value_col < df.shape[1]:
        if not _looks_like_numeric_series(df.iloc[first_data_row:, value_col], min_points=5):
            # Stop at separator after parsing at least one value column; otherwise try the next column.
            if parsed:
                break
            value_col += 1
            continue
        if not _looks_like_date_series(df.iloc[first_data_row:, date_col], min_points=5):
            break
        header = _first_non_empty_above(df, first_data_row, value_col)
        if _norm(header) in GENERIC_META_LABELS:
            header = _first_non_empty_above(df, header_row0, value_col)
        if not header:
            header = _first_non_empty_above(df, first_data_row, date_col)
        ric = ric_list[ric_index] if ric_index < len(ric_list) else ""
        display = header or ric or f"{sheet} {value_col + 1}"
        parsed.append(
            ParsedSeries(
                sheet=sheet,
                display_name=display,
                short_name=display,
                name_native=display,
                ric=ric,
                source=source,
                is_derived=False,
                date_col=date_col,
                value_col=value_col,
                header_row=header_row0,
                first_data_row=first_data_row,
            )
        )
        value_col += 1
        ric_index += 1
        if ric_list and ric_index >= len(ric_list):
            break
    return parsed


def _load_catalog_overrides(catalog_path: str | Path | None) -> dict[str, dict[str, Any]]:
    if not catalog_path:
        return {}
    path = Path(catalog_path)
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as f:
        data = yaml.safe_load(f) or {}
    rows = data.get("series", data if isinstance(data, list) else [])
    return {str(row.get("series_id")): dict(row) for row in rows if row.get("series_id")}


def _apply_catalog_overrides(df: pd.DataFrame, catalog_path: str | Path | None) -> pd.DataFrame:
    if not df.empty:
        if "term_type" not in df.columns:
            df = df.copy()
            df["term_type"] = "continuous"
        if "calendar_month" not in df.columns:
            df = df.copy()
            df["calendar_month"] = ""
    overrides = _load_catalog_overrides(catalog_path)
    if df.empty:
        return df
    if not overrides:
        return _apply_unit_rule_metadata(df)
    out = df.copy()
    editable_cols = [
        "display_name",
        "sector",
        "product",
        "region",
        "contract_month",
        "term_type",
        "calendar_month",
        "unit_native",
        "unit_normalized",
        "unit_conversion",
        "unit_source",
        "ric",
    ]
    override_frame = pd.DataFrame.from_dict(overrides, orient="index")
    if override_frame.empty:
        return out
    override_frame.index = override_frame.index.astype(str)
    series_ids = out["series_id"].astype(str)
    for col in editable_cols:
        if col not in out.columns or col not in override_frame.columns:
            continue
        mapped = series_ids.map(override_frame[col])
        mask = mapped.notna() & mapped.ne("")
        if mask.any():
            out.loc[mask, col] = mapped[mask].astype(str).to_numpy()
    return _apply_unit_rule_metadata(out)


def _ensure_runtime_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    out = df
    missing_defaults = {
        "term_type": "continuous",
        "calendar_month": "",
    }
    for col, default in missing_defaults.items():
        if col not in out.columns:
            out = out.copy()
            out[col] = default
    return _apply_unit_rule_metadata(out)


def _cache_write_frame(df: pd.DataFrame) -> pd.DataFrame:
    string_cols = [
        "series_id",
        "display_name",
        "sheet",
        "sector",
        "product",
        "region",
        "contract_month",
        "term_type",
        "calendar_month",
        "unit_native",
        "unit_normalized",
        "unit_conversion",
        "unit_source",
        "ric",
        "source",
        "name_native",
        "short_name",
    ]
    out = df.copy()
    for col in string_cols:
        if col in out.columns:
            out[col] = out[col].fillna("").astype(str)
    return out


def generate_catalog(df: pd.DataFrame, catalog_path: str | Path | None = None, overwrite: bool = False) -> Path:
    path = Path(catalog_path) if catalog_path else default_catalog_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    existing = _load_catalog_overrides(path)
    rows: list[dict[str, Any]] = []
    for _, row in (
        df.sort_values(["sheet", "sector", "product", "display_name"])
        .drop_duplicates("series_id")
        .iterrows()
    ):
        series_id = str(row["series_id"])
        base = {
            "series_id": series_id,
            "display_name": row.get("display_name", ""),
            "sheet": row.get("sheet", ""),
            "sector": row.get("sector", ""),
            "product": row.get("product", ""),
            "region": row.get("region", ""),
            "contract_month": row.get("contract_month", ""),
            "term_type": row.get("term_type", ""),
            "calendar_month": row.get("calendar_month", ""),
            "unit_native": row.get("unit_native", ""),
            "unit_normalized": row.get("unit_normalized", ""),
            "unit_conversion": row.get("unit_conversion", ""),
            "unit_source": row.get("unit_source", ""),
            "ric": row.get("ric", ""),
            "is_derived": bool(row.get("is_derived", False)),
            "source": row.get("source", ""),
        }
        if not overwrite and series_id in existing:
            base.update({key: value for key, value in existing[series_id].items() if value not in (None, "")})
        unit_rule = _unit_rule(str(base.get("sheet", "")), str(base.get("display_name", "")), str(base.get("ric", "")))
        if unit_rule.get("unit_native"):
            base["unit_native"] = unit_rule["unit_native"]
            base["unit_normalized"] = unit_rule["unit_normalized"]
            base["unit_conversion"] = unit_rule["conversion"]
            base["unit_source"] = unit_rule["source"]
        rows.append(base)
    with path.open("w", encoding="utf-8") as f:
        yaml.safe_dump({"series": rows}, f, allow_unicode=True, sort_keys=False, width=120)
    return path


def parse_nap_workbook(
    workbook_path: str | Path = DEFAULT_NAP_WORKBOOK,
    catalog_path: str | Path | None = None,
    generate_catalog_file: bool = True,
    sheets: list[str] | None = None,
) -> pd.DataFrame:
    workbook = Path(workbook_path)
    if not workbook.exists():
        raise FileNotFoundError(f"NAP workbook not found: {workbook}")

    wb_formula = load_workbook(workbook, read_only=True, data_only=False)
    wb_values = load_workbook(workbook, read_only=True, data_only=True)
    target_sheets = sheets or [sheet for sheet in RELEVANT_SHEETS if sheet in wb_values.sheetnames]
    all_records: list[dict[str, Any]] = []

    try:
        for sheet in target_sheets:
            if sheet not in wb_values.sheetnames:
                logger.warning("Sheet not found in NAP workbook: %s", sheet)
                continue
            values_ws = wb_values[sheet]
            formula_ws = wb_formula[sheet]
            if values_ws.max_row <= 1 and values_ws.max_column <= 1:
                continue
            df = _worksheet_to_frame(values_ws)
            formulas = _top_reuters_formulas(formula_ws)
            if not formulas:
                continue
            parsed_series: list[ParsedSeries] = []
            rdp_formulas = [item for item in formulas if item[2] == "RDP.HistoricalPricing"]
            rhistory_formulas = [item for item in formulas if item[2] == "RHistory"]
            if rdp_formulas:
                parsed_series.extend(_parse_rdp_sheet(sheet, df, rdp_formulas[0]))
            for formula in rhistory_formulas:
                parsed_series.extend(_parse_rhistory_formula_group(sheet, df, formula))

            for parsed in parsed_series:
                all_records.extend(_series_to_records(df, parsed))
            logger.info("Parsed %s NAP series from sheet %s", len(parsed_series), sheet)
    finally:
        wb_formula.close()
        wb_values.close()

    if not all_records:
        return pd.DataFrame(columns=STANDARD_COLUMNS + EXTRA_COLUMNS)

    out = pd.DataFrame(all_records)
    calendar_records = _derive_calendar_month_records(out)
    if calendar_records:
        out = pd.concat([out, pd.DataFrame(calendar_records)], ignore_index=True)
    out = out.sort_values(["series_id", "date"]).drop_duplicates(["date", "series_id"], keep="last")
    out = out[STANDARD_COLUMNS + [col for col in EXTRA_COLUMNS if col in out.columns]]

    if generate_catalog_file:
        generate_catalog(out, catalog_path or default_catalog_path(), overwrite=False)
    out = _apply_catalog_overrides(out, catalog_path or default_catalog_path())
    return out


def _write_cache(df: pd.DataFrame, cache_path: Path) -> None:
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    frame = _cache_write_frame(df)
    try:
        frame.to_parquet(cache_path, index=False)
    except Exception as exc:
        fallback = cache_path.with_suffix(cache_path.suffix + ".pkl")
        logger.warning("Unable to write parquet cache (%s); wrote pickle fallback: %s", exc, fallback)
        frame.to_pickle(fallback)


def _read_cache(cache_path: Path) -> pd.DataFrame | None:
    if cache_path.exists():
        try:
            return pd.read_parquet(cache_path)
        except Exception as exc:
            logger.warning("Unable to read parquet cache %s: %s", cache_path, exc)
    fallback = cache_path.with_suffix(cache_path.suffix + ".pkl")
    if fallback.exists():
        return pd.read_pickle(fallback)
    return None


def load_nap_timeseries(
    workbook_path: str | Path = DEFAULT_NAP_WORKBOOK,
    cache_path: str | Path | None = None,
    catalog_path: str | Path | None = None,
    refresh: bool = False,
) -> pd.DataFrame:
    workbook = Path(workbook_path)
    cache = Path(cache_path) if cache_path else default_cache_path()
    catalog = Path(catalog_path) if catalog_path else default_catalog_path()
    if not refresh:
        cached = _read_cache(cache)
        if cached is not None:
            fallback = cache.with_suffix(cache.suffix + ".pkl")
            cache_for_mtime = cache if cache.exists() else fallback
            if not workbook.exists() or cache_for_mtime.stat().st_mtime >= workbook.stat().st_mtime:
                cached = _ensure_runtime_columns(cached)
                if catalog.exists() and catalog.stat().st_mtime > cache_for_mtime.stat().st_mtime:
                    return _apply_catalog_overrides(cached, catalog)
                return cached

    df = parse_nap_workbook(workbook, catalog_path=catalog, generate_catalog_file=True)
    _write_cache(df, cache)
    return df


def series_catalog_frame(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    columns = [
        "series_id",
        "display_name",
        "sheet",
        "sector",
        "product",
        "region",
        "contract_month",
        "term_type",
        "calendar_month",
        "unit_native",
        "unit_normalized",
        "unit_conversion",
        "unit_source",
        "ric",
        "is_derived",
        "source",
        "name_native",
        "short_name",
    ]
    return df.sort_values(["sector", "product", "display_name"]).drop_duplicates("series_id")[columns]


def workbook_status(df: pd.DataFrame, workbook_path: str | Path = DEFAULT_NAP_WORKBOOK) -> dict[str, Any]:
    workbook = Path(workbook_path)
    latest = pd.to_datetime(df["date"]).max() if not df.empty else pd.NaT
    return {
        "workbook_path": str(workbook),
        "workbook_mtime": pd.Timestamp.fromtimestamp(workbook.stat().st_mtime) if workbook.exists() else pd.NaT,
        "latest_trade_date": latest,
        "series_count": int(df["series_id"].nunique()) if not df.empty else 0,
        "row_count": int(len(df)),
    }


def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(description="Parse Reuters NAP workbook into a normalized long table.")
    parser.add_argument("--workbook", default=str(DEFAULT_NAP_WORKBOOK))
    parser.add_argument("--cache", default=str(default_cache_path()))
    parser.add_argument("--catalog", default=str(default_catalog_path()))
    parser.add_argument("--refresh", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")
    df = load_nap_timeseries(args.workbook, args.cache, args.catalog, refresh=args.refresh)
    status = workbook_status(df, args.workbook)
    print(
        f"Parsed {status['series_count']} series / {status['row_count']:,} rows. "
        f"Latest date: {status['latest_trade_date']:%Y-%m-%d}"
    )


if __name__ == "__main__":
    main()
