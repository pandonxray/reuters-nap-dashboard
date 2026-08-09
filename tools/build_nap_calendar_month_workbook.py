from __future__ import annotations

import argparse
import json
import re
import sys
import types
from copy import copy
from pathlib import Path

if "yaml" not in sys.modules:
    sys.modules["yaml"] = types.SimpleNamespace(
        safe_load=lambda *args, **kwargs: {},
        safe_dump=lambda *args, **kwargs: None,
    )

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.dimensions import ColumnDimension

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src import nap_adapter as na  # noqa: E402


TARGET_SHEETS = [
    "Gasoline",
    "Heating Oil&Jet fuel",
    "Diesel",
    "Nap",
    "LNG",
    "Crk",
    "Margin",
    "Propane",
    "Fuel oil",
]

OUTPUT_SHEET_NAMES = {
    "Gasoline": "Gasoline_自然月",
    "Heating Oil&Jet fuel": "HOJet_自然月",
    "Diesel": "Diesel_自然月",
    "Nap": "Nap_自然月",
    "LNG": "LNG_自然月",
    "Crk": "Crk_自然月",
    "Margin": "Margin_自然月",
    "Propane": "Propane_自然月",
    "Fuel oil": "FuelOil_自然月",
}

LABEL_ALIASES = {
    "RB": "RBOB",
    "EBOBNWE": "EBOB",
    "MOG92SG": "新加坡92汽油纸货",
    "MOG95SG": "新加坡95汽油纸货",
    "HO": "HO",
    "JETFUSGC": "Jet USG",
    "JETFCNWE": "Jet NWE",
    "JETSG": "Jet Singapore",
    "LGO": "LSGO",
    "GO10SG": "新加坡10ppm柴油纸货",
    "NACFRJP": "MOPJ",
    "NAPJPEW": "石脑油东西价差",
    "NACFRJPCK": "MOPJ裂差",
    "NAPCNWEA": "NWE Naphtha",
    "NAPCNWEAC": "NWE Naphtha Crack",
    "A7Q": "美国天然气",
    "MOG92SGCK": "新加坡92汽油裂差",
    "EBOBNWECK": "EBOB裂差",
    "RBCCLC": "RBOB裂差",
    "GO10BRTCK": "新加坡柴油裂差",
    "LGOC": "欧洲柴油裂差",
    "HOCCLC": "HO裂差",
    "JETSGCK": "新加坡航煤裂差",
    "JETFCNWECK": "欧洲航煤裂差",
    "FO380BRTCK": "新加坡高硫裂差",
    "HFOFARAAC": "欧洲高硫裂差",
    "PROCNWE": "NWE LPG",
    "PROFEI": "FEI",
    "FO380SG": "新加坡高硫燃料油",
    "HFOFARAA": "欧洲高硫燃料油",
}

SHEET_PALETTE = [
    "D9EAF7",
    "E3F2DE",
    "FFF1CC",
    "F5DDE8",
    "E6E0F8",
    "DCEDEA",
    "FBE1D0",
    "E7EAF0",
]


def default_input_path() -> Path:
    base = Path.home() / "Nutstore" / "1"
    candidates = list(base.rglob("Nap.xlsx"))
    if not candidates:
        raise FileNotFoundError("未在用户 Nutstore 目录下找到 Nap.xlsx，请用 --input 指定路径。")
    return max(candidates, key=lambda path: path.stat().st_mtime)


def base_from_ric(ric: str) -> str:
    value = (ric or "").upper()
    value = re.sub(r"(?:SW)?MC\d{1,2}$", "", value)
    if re.search(r"[-=/]", value):
        value = re.sub(r"C\d{1,2}(?=$|[-=/])", "C", value)
    else:
        value = re.sub(r"C\d{1,2}$", "", value)
    return re.sub(r"[^A-Z0-9]+", "", value)


def base_label(display_name: str, ricbase: str) -> str:
    if ricbase in LABEL_ALIASES:
        return LABEL_ALIASES[ricbase]
    label = display_name or ricbase
    patterns = [
        r"\b(monthly\s+)?continuation\s*\d{1,2}\b",
        r"\bmonth\s+continuation\s*\d{1,2}\b",
        r"M\s*\d{1,2}$",
        r"c\s*\d{1,2}$",
        r"连\s*\d{1,2}$",
    ]
    for pattern in patterns:
        label = re.sub(pattern, "", label, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", label).strip(" -_/") or ricbase


def parse_groups(workbook_path: Path) -> dict[str, list[dict]]:
    wb_formula = load_workbook(workbook_path, read_only=True, data_only=False)
    wb_values = load_workbook(workbook_path, read_only=True, data_only=True)
    groups_by_sheet: dict[str, list[dict]] = {}
    try:
        for sheet in TARGET_SHEETS:
            if sheet not in wb_values.sheetnames:
                groups_by_sheet[sheet] = []
                continue

            df = na._worksheet_to_frame(wb_values[sheet])
            formulas = na._top_reuters_formulas(wb_formula[sheet])
            parsed = []
            rdp = [item for item in formulas if item[2] == "RDP.HistoricalPricing"]
            rhistory = [item for item in formulas if item[2] == "RHistory"]
            if rdp:
                parsed.extend(na._parse_rdp_sheet(sheet, df, rdp[0]))
            for formula in rhistory:
                parsed.extend(na._parse_rhistory_formula_group(sheet, df, formula))

            raw_groups: dict[str, dict] = {}
            for item in parsed:
                display_name = item.display_name or item.short_name or item.name_native or item.ric
                contract_month = na._infer_contract_month(display_name, item.ric)
                match = re.match(r"^M(\d{1,2})$", contract_month or "")
                if not match:
                    continue
                month_no = int(match.group(1))
                if month_no < 1 or month_no > 12:
                    continue
                ricbase = base_from_ric(item.ric)
                if not ricbase:
                    continue
                raw_groups.setdefault(
                    ricbase,
                    {
                        "label": base_label(display_name, ricbase),
                        "ricbase": ricbase,
                        "source_sheet": sheet,
                        "months": {},
                    },
                )["months"][month_no] = item

            complete = []
            for group in raw_groups.values():
                if set(group["months"]) != set(range(1, 13)):
                    continue
                group["month_items"] = [group["months"][month] for month in range(1, 13)]
                del group["months"]
                complete.append(group)
            groups_by_sheet[sheet] = complete
    finally:
        wb_formula.close()
        wb_values.close()
    return groups_by_sheet


def copy_dimensions(src_ws, dst_ws, max_col: int) -> None:
    for idx in range(1, max_col + 1):
        letter = get_column_letter(idx)
        dim = ColumnDimension(dst_ws, index=letter)
        if idx == 1:
            dim.width = 13
        else:
            dim.width = 12
        dst_ws.column_dimensions[letter] = dim


def style_sheet(ws, max_row: int, max_col: int, groups: list[dict]) -> None:
    title_fill = PatternFill("solid", fgColor="16324F")
    note_fill = PatternFill("solid", fgColor="EEF3F7")
    month_fill = PatternFill("solid", fgColor="F5F7FA")
    thin = Side(style="thin", color="CDD6E0")
    border = Border(bottom=thin, right=thin)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B6"
    ws.auto_filter.ref = f"A5:{get_column_letter(max_col)}{max_row}"

    ws["A1"].fill = title_fill
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 26

    for row in (2, 3):
        for cell in ws[row]:
            cell.fill = note_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    for cell in ws[4]:
        cell.font = Font(bold=True, color="1F2937")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for cell in ws[5]:
        cell.fill = month_fill
        cell.font = Font(bold=True, color="1F2937")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws["A5"].fill = PatternFill("solid", fgColor="DDE7F0")
    ws["A5"].number_format = "@"

    for row in range(6, max_row + 1):
        ws.cell(row=row, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    for col in range(2, max_col + 1):
        for row in range(6, max_row + 1):
            ws.cell(row=row, column=col).number_format = "#,##0.00"

    for group_idx, _group in enumerate(groups):
        start_col = 2 + group_idx * 12
        end_col = start_col + 11
        fill = PatternFill("solid", fgColor=SHEET_PALETTE[group_idx % len(SHEET_PALETTE)])
        for col in range(start_col, end_col + 1):
            ws.cell(row=4, column=col).fill = fill
            ws.cell(row=5, column=col).fill = fill


def source_range(item, attr: str, extra_rows: int) -> str:
    col_idx = item.date_col + 1 if attr == "date" else item.value_col + 1
    col = get_column_letter(col_idx)
    start_row = item.first_data_row + 1
    end_row = start_row + extra_rows - 1
    quoted = quote_sheetname(item.sheet)
    return f"{quoted}!${col}${start_row}:${col}${end_row}"


def dynamic_source_range(item, attr: str, source_row_limit: int) -> str:
    col_idx = item.date_col + 1 if attr == "date" else item.value_col + 1
    col = get_column_letter(col_idx)
    start_row = item.first_data_row + 1
    quoted = quote_sheetname(item.sheet)
    return f"{quoted}!${col}${start_row}:${col}${source_row_limit}"


def source_cell_ref(item, attr: str, row_offset: int) -> str:
    col_idx = item.date_col + 1 if attr == "date" else item.value_col + 1
    col = get_column_letter(col_idx)
    row = item.first_data_row + 1 + row_offset
    quoted = quote_sheetname(item.sheet)
    return f"{quoted}!${col}{row}"


def source_value_matrix_range(group: dict, extra_rows: int) -> str:
    month_items = group["month_items"]
    first = month_items[0]
    last = month_items[-1]
    start_col = get_column_letter(first.value_col + 1)
    end_col = get_column_letter(last.value_col + 1)
    start_row = first.first_data_row + 1
    end_row = start_row + extra_rows - 1
    quoted = quote_sheetname(first.sheet)
    return f"{quoted}!${start_col}${start_row}:${end_col}${end_row}"


def dynamic_value_matrix_range(group: dict, source_row_limit: int) -> str:
    month_items = group["month_items"]
    first = month_items[0]
    last = month_items[-1]
    start_col = get_column_letter(first.value_col + 1)
    end_col = get_column_letter(last.value_col + 1)
    start_row = first.first_data_row + 1
    quoted = quote_sheetname(first.sheet)
    return f"{quoted}!${start_col}${start_row}:${end_col}${source_row_limit}"


def add_calendar_sheet(
    wb,
    source_sheet: str,
    groups: list[dict],
    source_row_limit: int,
    formula_entries: list[dict[str, str]] | None = None,
) -> None:
    output_name = OUTPUT_SHEET_NAMES[source_sheet]
    if output_name in wb.sheetnames:
        del wb[output_name]
    ws = wb.create_sheet(output_name)

    if not groups:
        ws["A1"] = f"{source_sheet}：C1-C12 反算自然月"
        ws["A2"] = "未发现完整的 C1-C12 连续月产品组；因此这里只保留说明，不生成自然月公式列。"
        style_sheet(ws, 2, 1, [])
        return

    max_col = 1 + len(groups) * 12
    max_row = 6

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws["A1"] = f"{source_sheet}：C1-C12 连续月反算为 1-12 月自然月"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws["A2"] = (
        "公式逻辑：每一行先看日期属于哪个月份。C1 等于该日期所在月份，C2 等于下一个月份；"
        "目标自然月 m 使用 Ck，k=MOD(m-MONTH(日期),12)+1。比如日期在 7 月，7 月取 C1，8 月取 C2，次年 6 月取 C12。"
    )
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)
    ws["A3"] = (
        "输出说明：只生成完整 C1-C12 的产品；单一现货价或缺月品种不会展开。"
        "第 6 行是 Excel 365 动态数组公式，Reuters 刷新后会随原始日期列自动向下展开并重算。"
    )

    ws["A4"] = "日期"
    ws["A5"] = "日期"

    for group_idx, group in enumerate(groups):
        start_col = 2 + group_idx * 12
        end_col = start_col + 11
        ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=end_col)
        ws.cell(row=4, column=start_col).value = group["label"]
        for month in range(1, 13):
            cell = ws.cell(row=5, column=start_col + month - 1)
            cell.value = month
            cell.number_format = '0"月"'
            cell.comment = Comment(
                f"{group['label']} 自然月 {month} 月。根据本行日期自动选择 C1-C12 中对应的连续月。", "Codex"
            )

    first_date_item = groups[0]["month_items"][0]
    first_date_range = dynamic_source_range(first_date_item, "date", source_row_limit)
    date_formula = f'=FILTER({first_date_range},{first_date_range}<>"")'
    if formula_entries is None:
        ws.cell(row=6, column=1).value = date_formula
    else:
        formula_entries.append({"sheet": output_name, "cell": "A6", "formula": date_formula})
    ws.cell(row=6, column=1).number_format = "yyyy-mm-dd"

    for group_idx, group in enumerate(groups):
        col = 2 + group_idx * 12
        value_matrix = dynamic_value_matrix_range(group, source_row_limit)
        formula = (
            f'=LET(d,$A$6#,src,{value_matrix},'
            'MAKEARRAY(ROWS(d),12,LAMBDA(r,m,LET('
            'c,1+2*MOD(m-MONTH(INDEX(d,r)),12),'
            'IFERROR(IF(COUNTBLANK(INDEX(src,r,c))>0,"",INDEX(src,r,c)),"")))))'
        )
        if formula_entries is None:
            ws.cell(row=6, column=col).value = formula
        else:
            formula_entries.append(
                {
                    "sheet": output_name,
                    "cell": f"{get_column_letter(col)}6",
                    "formula": formula,
                }
            )
        ws.cell(row=6, column=col).number_format = "#,##0.00"

    copy_dimensions(wb[source_sheet], ws, max_col)
    style_sheet(ws, max_row, max_col, groups)


def add_filled_calendar_sheet(wb, source_sheet: str, groups: list[dict], future_rows: int) -> None:
    output_name = OUTPUT_SHEET_NAMES[source_sheet]
    if output_name in wb.sheetnames:
        del wb[output_name]
    ws = wb.create_sheet(output_name)

    if not groups:
        ws["A1"] = f"{source_sheet}：C1-C12 反算自然月"
        ws["A2"] = "未发现完整的 C1-C12 连续月产品组；因此这里只保留说明，不生成自然月公式列。"
        style_sheet(ws, 2, 1, [])
        return

    source_ws = wb[source_sheet]
    data_rows = max(1, source_ws.max_row - 1 + future_rows)
    max_row = 5 + data_rows
    max_col = 1 + len(groups) * 12

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws["A1"] = f"{source_sheet}：C1-C12 连续月反算为 1-12 月自然月"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws["A2"] = (
        "公式逻辑：每一行先看日期属于哪个月份。C1 等于该日期所在月份，C2 等于下一个月份；"
        "目标自然月 m 使用 Ck，k=MOD(m-MONTH(日期),12)+1。比如日期在 7 月，7 月取 C1，8 月取 C2，次年 6 月取 C12。"
    )
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)
    ws["A3"] = (
        f"输出说明：只生成完整 C1-C12 的产品；单一现货价或缺月品种不会展开。"
        f"本页已向下预填 {future_rows} 行未来公式；Reuters 刷新源 sheet 后，这些公式会自动读取新增行。"
    )

    ws["A4"] = "日期"
    ws["A5"] = "日期"

    for group_idx, group in enumerate(groups):
        start_col = 2 + group_idx * 12
        end_col = start_col + 11
        ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=end_col)
        ws.cell(row=4, column=start_col).value = group["label"]
        for month in range(1, 13):
            cell = ws.cell(row=5, column=start_col + month - 1)
            cell.value = month
            cell.number_format = '0"月"'
            cell.comment = Comment(
                f"{group['label']} 自然月 {month} 月。根据本行日期自动选择 C1-C12 中对应的连续月。", "Codex"
            )

    first_date_item = groups[0]["month_items"][0]
    first_date_range = source_range(first_date_item, "date", data_rows)
    for row in range(6, max_row + 1):
        ws.cell(row=row, column=1).value = f'=IFERROR(INDEX({first_date_range},ROW()-5),"")'

    for group_idx, group in enumerate(groups):
        value_ranges = [source_range(item, "value", data_rows) for item in group["month_items"]]
        for target_month in range(1, 13):
            col = 2 + group_idx * 12 + target_month - 1
            month_header = f"{get_column_letter(col)}$5"
            choices = ",".join([f"INDEX({rng},ROW()-5)" for rng in value_ranges])
            for row in range(6, max_row + 1):
                ws.cell(row=row, column=col).value = (
                    f'=IF($A{row}="","",IFERROR(CHOOSE(MOD({month_header}-MONTH($A{row}),12)+1,'
                    f'{choices}),""))'
                )

    copy_dimensions(source_ws, ws, max_col)
    style_sheet(ws, max_row, max_col, groups)


def add_light_calendar_sheet(wb, source_sheet: str, groups: list[dict], future_rows: int) -> None:
    output_name = OUTPUT_SHEET_NAMES[source_sheet]
    if output_name in wb.sheetnames:
        del wb[output_name]
    ws = wb.create_sheet(output_name)

    if not groups:
        ws["A1"] = f"{source_sheet}：C1-C12 反算自然月"
        ws["A2"] = "未发现完整的 C1-C12 连续月产品组；因此这里只保留说明，不生成自然月公式列。"
        style_sheet(ws, 2, 1, [])
        return

    source_ws = wb[source_sheet]
    data_rows = max(1, source_ws.max_row - 1 + future_rows)
    max_row = 5 + data_rows
    max_col = 1 + len(groups) * 12

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws["A1"] = f"{source_sheet}：C1-C12 连续月反算为 1-12 月自然月（轻公式版）"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws["A2"] = (
        "轻公式逻辑：每个值单元格只用一个 INDEX 在 C1-C12 横向矩阵中取数。"
        "目标自然月 m 使用 Ck，k=MOD(m-MONTH(日期),12)+1；例如日期在 7 月，7 月取 C1，8 月取 C2，次年 6 月取 C12。"
    )
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)
    ws["A3"] = (
        f"本页已向下预填 {future_rows} 行未来公式；Reuters 刷新源 sheet 后，这些公式会自动读取新增行。"
        "相比旧版 CHOOSE+12个INDEX，轻公式版显著减少重算依赖。"
    )

    ws["A4"] = "日期"
    ws["A5"] = "日期"

    for group_idx, group in enumerate(groups):
        start_col = 2 + group_idx * 12
        end_col = start_col + 11
        ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=end_col)
        ws.cell(row=4, column=start_col).value = group["label"]
        for month in range(1, 13):
            cell = ws.cell(row=5, column=start_col + month - 1)
            cell.value = month
            cell.number_format = '0"月"'
            cell.comment = Comment(
                f"{group['label']} 自然月 {month} 月。根据本行日期自动选择 C1-C12 中对应的连续月。", "Codex"
            )

    first_date_item = groups[0]["month_items"][0]
    for row in range(6, max_row + 1):
        row_offset = row - 6
        date_ref = source_cell_ref(first_date_item, "date", row_offset)
        ws.cell(row=row, column=1).value = f'=IF({date_ref}="","",{date_ref})'

    for group_idx, group in enumerate(groups):
        value_matrix = source_value_matrix_range(group, data_rows)
        for target_month in range(1, 13):
            col = 2 + group_idx * 12 + target_month - 1
            month_header = f"{get_column_letter(col)}$5"
            for row in range(6, max_row + 1):
                ws.cell(row=row, column=col).value = (
                    f'=IF($A{row}="","",IFERROR(INDEX({value_matrix},ROW()-5,'
                    f'1+2*MOD({month_header}-MONTH($A{row}),12)),""))'
                )

    copy_dimensions(source_ws, ws, max_col)
    style_sheet(ws, max_row, max_col, groups)


def build_workbook(
    input_path: Path,
    output_path: Path,
    source_row_limit: int,
    formula_manifest_path: Path | None = None,
    formula_mode: str = "filled",
    future_rows: int = 1500,
) -> dict[str, int]:
    groups_by_sheet = parse_groups(input_path)
    wb = load_workbook(input_path, data_only=False, keep_links=True)
    formula_entries: list[dict[str, str]] | None = [] if formula_manifest_path else None
    try:
        for sheet in TARGET_SHEETS:
            if formula_mode == "dynamic":
                add_calendar_sheet(wb, sheet, groups_by_sheet.get(sheet, []), source_row_limit, formula_entries)
            elif formula_mode == "light":
                add_light_calendar_sheet(wb, sheet, groups_by_sheet.get(sheet, []), future_rows)
            else:
                add_filled_calendar_sheet(wb, sheet, groups_by_sheet.get(sheet, []), future_rows)

        if hasattr(wb, "calculation"):
            wb.calculation.fullCalcOnLoad = formula_mode != "light"
            wb.calculation.forceFullCalc = formula_mode != "light"
            wb.calculation.calcMode = "auto"

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        if formula_manifest_path and formula_entries is not None:
            formula_manifest_path.parent.mkdir(parents=True, exist_ok=True)
            formula_manifest_path.write_text(
                json.dumps(formula_entries, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
    finally:
        wb.close()
    return {sheet: len(groups) for sheet, groups in groups_by_sheet.items()}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, default=None)
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument("--source-row-limit", type=int, default=20000)
    parser.add_argument("--formula-manifest", type=Path, default=None)
    parser.add_argument("--formula-mode", choices=["filled", "dynamic", "light"], default="filled")
    parser.add_argument("--future-rows", type=int, default=1500)
    args = parser.parse_args()

    input_path = args.input or default_input_path()
    output_path = args.output or (ROOT / "outputs" / "Nap_自然月反算.xlsx")
    summary = build_workbook(
        input_path,
        output_path,
        args.source_row_limit,
        args.formula_manifest,
        args.formula_mode,
        args.future_rows,
    )
    print(f"OUTPUT={output_path}")
    for sheet, count in summary.items():
        print(f"{sheet}: {count} full C1-C12 groups")


if __name__ == "__main__":
    main()
